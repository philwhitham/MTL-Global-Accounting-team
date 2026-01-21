#!/usr/bin/env python3
"""
Simple script to read Excel files using openpyxl
"""
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook

def read_excel_file(filepath):
    """Read an Excel file and print its contents"""
    try:
        wb = load_workbook(filepath, data_only=True)
        print(f"\n📊 Excel File: {filepath}")
        print(f"📋 Number of sheets: {len(wb.sheetnames)}")
        print(f"📝 Sheet names: {wb.sheetnames}\n")
        
        for sheet_name in wb.sheetnames:
            print(f"\n{'='*100}")
            print(f"SHEET: {sheet_name}")
            print(f"{'='*100}\n")
            
            ws = wb[sheet_name]
            
            # Print first 100 rows
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx > 100:
                    print(f"\n... (showing first 100 rows only)\n")
                    break
                # Filter out completely empty rows
                if any(cell is not None and str(cell).strip() != '' for cell in row):
                    row_str = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    print(f"Row {row_idx}: {row_str}")
            
            print("\n")
            
    except Exception as e:
        print(f"Error reading file: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 read_excel_simple.py <excel_file>")
        sys.exit(1)
    
    read_excel_file(sys.argv[1])
